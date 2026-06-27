import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import re
import shutil
import tempfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.styles import Font


# ── 辅助：.xls → .xlsx 转换 ──────────────────────────────────────
def convert_xls_to_xlsx(src_path):
    """若源文件为 .xls，调用 Excel COM 转换为临时 .xlsx（保留完整格式）。
       返回 (xlsx_path, 是否为临时文件)。"""
    if not src_path.lower().endswith('.xls'):
        return src_path, False

    try:
        import win32com.client
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(src_path))
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()
        wb.SaveAs(tmp_path, FileFormat=51)
        wb.Close(SaveChanges=False)
        excel.Quit()
        return tmp_path, True
    except Exception:
        pass

    # 回退：xlrd
    try:
        import xlrd
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()
        rb = xlrd.open_workbook(src_path, formatting_info=True)
        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        for sheet_name in rb.sheet_names():
            df = pd.read_excel(src_path, sheet_name=sheet_name, header=None)
            ws = wb_out.create_sheet(title=sheet_name)
            for r_idx, row in df.iterrows():
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx + 1, column=c_idx, value=val)
        wb_out.save(tmp_path)
        return tmp_path, True
    except Exception as e2:
        raise RuntimeError(f'无法转换 .xls 文件: {e2}')


# ── 辅助：跨 workbook 复制 sheet ─────────────────────────────────
def copy_sheet_across_workbooks(wb_src, wb_dst, sheet_name):
    """把 wb_src 中的 sheet 完整复制到 wb_dst 中（值、格式、合并单元格、行列尺寸）"""
    ws_src = wb_src[sheet_name]

    # 若目标已存在同名 sheet，先删除
    if sheet_name in wb_dst.sheetnames:
        del wb_dst[sheet_name]

    ws_dst = wb_dst.create_sheet(sheet_name)

    # 复制列宽
    for col_letter, col_dim in ws_src.column_dimensions.items():
        if col_dim.width:
            ws_dst.column_dimensions[col_letter].width = col_dim.width

    # 复制行高
    for row_num, row_dim in ws_src.row_dimensions.items():
        if row_dim.height:
            ws_dst.row_dimensions[row_num].height = row_dim.height

    # 复制合并单元格
    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))

    # 复制单元格值与格式（跳过 MergedCell）
    for row in ws_src.iter_rows():
        for cell in row:
            if type(cell).__name__ == 'MergedCell':
                continue
            new_cell = ws_dst[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    return ws_dst


# ── 核心逻辑函数 ───────────────────────────────────────────────────
def parse_orders(src_path):
    """解析送货单，返回 (orders, order_dates)"""
    df_raw = pd.read_excel(src_path, sheet_name='送货单', header=None)
    orders = {}
    order_dates = {}
    current_date = None

    for _, row in df_raw.iterrows():
        date_cell = str(row[9]) if pd.notna(row[9]) else ''
        m = re.search(r'送货日期：(\d{4}-\d{2}-\d{2})', date_cell)
        if m:
            date_str = m.group(1)
            day = int(date_str.split('-')[2])
            current_date = day
            orders.setdefault(current_date, [])
            order_dates[day] = date_str
            continue

        if current_date is None:
            continue

        seq  = row[1]
        name = row[3]
        if pd.notna(seq) and pd.notna(name):
            try:
                seq_int = int(float(str(seq)))
            except (ValueError, TypeError):
                continue
            name_s = str(name).strip()
            if seq_int > 0 and name_s not in ('品名', '序号') and '送货日期' not in name_s:
                unit   = str(row[7]).strip() if pd.notna(row[7]) else ''
                qty    = row[6] if pd.notna(row[6]) else 0
                price  = row[8] if pd.notna(row[8]) else 0
                amount = row[9] if pd.notna(row[9]) else 0
                note   = str(row[10]).strip() \
                         if pd.notna(row[10]) and str(row[10]).strip() not in ('nan', '') else ''
                orders[current_date].append((name_s, unit, qty, price, amount, note))

    return orders, order_dates


def adjust_and_fill_sheet(ws, items, day, full_date):
    """动态调整数据区为N行，填入数据，更新合计行"""
    n = len(items)

    # 取消所有涉及 Row >= 4 的合并单元格
    merges = [mc for mc in list(ws.merged_cells.ranges) if mc.min_row >= 4]
    for mc in merges:
        ws.unmerge_cells(str(mc))

    # 定位合计行
    total_row = None
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and '合计' in str(val):
            total_row = r
            break
    if total_row is None:
        total_row = ws.max_row

    current_data_rows = total_row - 4

    # 备份参考格式（数据区第一行）
    ref_fonts, ref_borders, ref_alignments, ref_num_fmts = {}, {}, {}, {}
    for c in range(1, 8):
        cell = ws.cell(row=4, column=c)
        ref_fonts[c]      = copy(cell.font)
        ref_borders[c]    = copy(cell.border)
        ref_alignments[c] = copy(cell.alignment)
        ref_num_fmts[c]   = cell.number_format

    total_font_a = copy(ws.cell(row=total_row, column=1).font)
    total_font_f = copy(ws.cell(row=total_row, column=6).font)
    total_num_fmt_f = ws.cell(row=total_row, column=6).number_format

    # 调整行数
    if n < current_data_rows:
        ws.delete_rows(4 + n, current_data_rows - n)
        new_total_row = 4 + n
    elif n > current_data_rows:
        rows_to_insert = n - current_data_rows
        ws.insert_rows(total_row, rows_to_insert)
        new_total_row = total_row + rows_to_insert
        for r in range(total_row, total_row + rows_to_insert):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.font      = ref_fonts[c]
                cell.border    = ref_borders[c]
                cell.alignment = ref_alignments[c]
                cell.number_format = ref_num_fmts[c]
            ws.cell(row=r, column=6).value = f'=D{r}*E{r}'
            ws.cell(row=r, column=6).number_format = ref_num_fmts[6]
    else:
        new_total_row = total_row

    # 填入数据
    for i, (name, unit, qty, price, amount, note) in enumerate(items):
        row_num = 4 + i
        ws.cell(row=row_num, column=1).value = full_date
        ws.cell(row=row_num, column=2).value = name
        ws.cell(row=row_num, column=3).value = unit
        ws.cell(row=row_num, column=4).value = qty
        ws.cell(row=row_num, column=5).value = price
        ws.cell(row=row_num, column=6).value = amount
        ws.cell(row=row_num, column=7).value = note if note else None
        for c in range(1, 8):
            cell = ws.cell(row=row_num, column=c)
            if cell.number_format == 'General':
                cell.number_format = ref_num_fmts[c]

    # 更新合计行
    ws.cell(row=new_total_row, column=1).value = '合计'
    ws.merge_cells(start_row=new_total_row, start_column=1,
                   end_row=new_total_row, end_column=5)
    ws.cell(row=new_total_row, column=6).value = f'=SUM(F4:F{3 + n})'
    ws.cell(row=new_total_row, column=6).number_format = total_num_fmt_f
    ws.cell(row=new_total_row, column=1).font = Font(
        name=total_font_a.name, size=total_font_a.size, bold=True,
        color=total_font_a.color, italic=total_font_a.italic,
        underline=total_font_a.underline, strike=total_font_a.strike
    )
    ws.cell(row=new_total_row, column=6).font = Font(
        name=total_font_f.name, size=total_font_f.size, bold=True,
        color=total_font_f.color, italic=total_font_f.italic,
        underline=total_font_f.underline, strike=total_font_f.strike
    )

    return new_total_row


def fill_ledger(template_path, bill_path, dst_path, log_func, progress_func):
    """主填充函数
       template_path: 台账模板文件
       bill_path:     对账单文件（含"送货单"sheet）
       dst_path:      输出文件路径
    """
    # 0. 复制模板为副本
    shutil.copy2(template_path, dst_path)
    log_func(f'✅ 已复制模板 → {dst_path}')

    # 1. 转换对账单（若为 .xls）
    bill_xlsx, is_tmp = convert_xls_to_xlsx(bill_path)
    if bill_path.lower().endswith('.xls'):
        log_func(f'[INFO] 已将 .xls 转换为 .xlsx 进行处理')

    # 2. 拷贝"送货单"到副本
    log_func('📋 正在拷贝"送货单"到副本...')
    wb_dst = load_workbook(dst_path)
    wb_bill = load_workbook(bill_xlsx)
    copy_sheet_across_workbooks(wb_bill, wb_dst, '送货单')
    wb_dst.save(dst_path)
    wb_bill.close()
    log_func(f'✅ "送货单"已拷贝到副本')

    # 3. 解析送货单
    log_func('📋 正在解析送货单数据...')
    orders, order_dates = parse_orders(bill_xlsx)
    has_data_days = set(orders.keys())
    log_func(f'✅ 共解析到 {len(orders)} 个送货日期: {sorted(has_data_days)}')

    # 4. 填充台账
    wb = load_workbook(dst_path)
    filled = 0
    total = len(has_data_days)

    for idx, day in enumerate(sorted(has_data_days), 1):
        progress_func(int(idx / total * 40))
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        items = orders[day]
        full_date = order_dates.get(day, f'2026-01-{day:02d}')
        adjust_and_fill_sheet(ws, items, day, full_date)

        filled += 1
        log_func(f'  📅 Sheet {sheet_name}: {len(items)} 条（日期 {full_date}），已动态调整表格')

    # 5. 汇总表处理
    log_func('📊 正在处理汇总表...')
    progress_func(50)
    ws_sum = wb['汇总']

    deleted_rows = 0
    for day in range(31, 0, -1):
        row_num = day + 1
        if day not in has_data_days:
            ws_sum.delete_rows(row_num, 1)
            deleted_rows += 1

    for r in range(2, ws_sum.max_row):
        date_val = ws_sum.cell(row=r, column=1).value
        if date_val and isinstance(date_val, str) and '.' in date_val:
            try:
                day = int(date_val.split('.')[-1])
                total_amt = sum(item[4] for item in orders[day])
                ws_sum.cell(row=r, column=2).value = total_amt
                if day in order_dates:
                    real_date = order_dates[day].replace('-', '.')
                    ws_sum.cell(row=r, column=1).value = real_date
            except (ValueError, IndexError, KeyError):
                pass

    last_data_row = ws_sum.max_row - 1
    ws_sum.cell(row=ws_sum.max_row, column=2).value = f'=SUM(B2:B{last_data_row})'
    log_func(f'  汇总表: 删除 {deleted_rows} 个空行，保留 {last_data_row - 1} 个日期')

    # 6. 删除无数据的日期sheet
    progress_func(80)
    log_func('🗑️ 正在删除无数据的日期表格...')
    deleted_sheets = 0
    for day in range(1, 32):
        sheet_name = str(day)
        if sheet_name in wb.sheetnames and day not in has_data_days:
            del wb[sheet_name]
            deleted_sheets += 1
    log_func(f'  删除 {deleted_sheets} 个空表格')

    wb.save(dst_path)
    if is_tmp:
        os.remove(bill_xlsx)
    progress_func(100)
    log_func(f'\n🎉 完成！')
    log_func(f'  填充台账表: {filled} 个')
    log_func(f'  汇总表保留: {last_data_row - 1} 个日期')
    log_func(f'  删除空表格: {deleted_sheets} 个')
    log_func(f'📁 输出文件: {dst_path}')
    return filled, deleted_sheets


# ── GUI 主体 ────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('饭堂菜肉台账自动填充工具')
        self.geometry('680x560')
        self.resizable(False, False)
        self.configure(bg='#f5f5f5')
        self._build_ui()

    def _build_ui(self):
        pad = {'padx': 16, 'pady': 5}

        # ── 标题 ──
        tk.Label(self, text='🍖 饭堂菜肉台账自动填充工具',
                 font=('微软雅黑', 14, 'bold'), bg='#f5f5f5', fg='#333').pack(pady=(16, 4))
        tk.Label(self, text='基于台账模板 + 对账单送货单，自动生成填充结果',
                 font=('微软雅黑', 9), bg='#f5f5f5', fg='#888').pack()

        ttk.Separator(self, orient='horizontal').pack(fill='x', padx=16, pady=8)

        # ── 模板文件 ──
        frame_tpl = tk.Frame(self, bg='#f5f5f5')
        frame_tpl.pack(fill='x', **pad)
        tk.Label(frame_tpl, text='台账模板：', width=14,
                 anchor='w', bg='#f5f5f5', font=('微软雅黑', 9)).pack(side='left')
        self.tpl_var = tk.StringVar(value='')
        tk.Entry(frame_tpl, textvariable=self.tpl_var, width=44,
                 font=('微软雅黑', 9)).pack(side='left', padx=(0, 6))
        tk.Button(frame_tpl, text='浏览', command=self._pick_tpl,
                  bg='#e0e0e0', relief='flat', padx=8).pack(side='left')

        # ── 对账单文件 ──
        frame_bill = tk.Frame(self, bg='#f5f5f5')
        frame_bill.pack(fill='x', **pad)
        tk.Label(frame_bill, text='对账单文件：', width=14,
                 anchor='w', bg='#f5f5f5', font=('微软雅黑', 9)).pack(side='left')
        self.bill_var = tk.StringVar()
        tk.Entry(frame_bill, textvariable=self.bill_var, width=44,
                 font=('微软雅黑', 9)).pack(side='left', padx=(0, 6))
        tk.Button(frame_bill, text='浏览', command=self._pick_bill,
                  bg='#e0e0e0', relief='flat', padx=8).pack(side='left')

        # ── 输出文件 ──
        frame_dst = tk.Frame(self, bg='#f5f5f5')
        frame_dst.pack(fill='x', **pad)
        tk.Label(frame_dst, text='输出文件路径：', width=14,
                 anchor='w', bg='#f5f5f5', font=('微软雅黑', 9)).pack(side='left')
        self.dst_var = tk.StringVar()
        tk.Entry(frame_dst, textvariable=self.dst_var, width=44,
                 font=('微软雅黑', 9)).pack(side='left', padx=(0, 6))
        tk.Button(frame_dst, text='浏览', command=self._pick_dst,
                  bg='#e0e0e0', relief='flat', padx=8).pack(side='left')

        # ── 运行按钮 ──
        self.run_btn = tk.Button(self, text='▶  开始填充', command=self._run,
                                 font=('微软雅黑', 11, 'bold'),
                                 bg='#4CAF50', fg='white',
                                 activebackground='#388E3C', activeforeground='white',
                                 relief='flat', padx=20, pady=6, cursor='hand2')
        self.run_btn.pack(pady=10)

        # ── 进度条 ──
        self.progress = ttk.Progressbar(self, length=640, mode='determinate')
        self.progress.pack(padx=16, pady=(0, 8))

        # ── 日志区域 ──
        tk.Label(self, text='执行日志：', anchor='w',
                 bg='#f5f5f5', font=('微软雅黑', 9)).pack(anchor='w', padx=16)
        frame_log = tk.Frame(self, bg='#f5f5f5')
        frame_log.pack(fill='both', expand=True, padx=16, pady=(0, 16))

        self.log_box = tk.Text(frame_log, height=12, font=('Consolas', 9),
                               bg='#1e1e1e', fg='#d4d4d4',
                               relief='flat', state='disabled', wrap='word')
        scroll = ttk.Scrollbar(frame_log, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=scroll.set)
        self.log_box.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')

    def _pick_tpl(self):
        path = filedialog.askopenfilename(
            title='选择台账模板文件',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm'), ('所有文件', '*.*')])
        if path:
            self.tpl_var.set(path)

    def _pick_bill(self):
        path = filedialog.askopenfilename(
            title='选择对账单文件',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xls'), ('所有文件', '*.*')])
        if path:
            self.bill_var.set(path)
            # 每次选择对账单后，自动重新推导输出路径
            bill_dir = os.path.dirname(path)
            bill_name = os.path.splitext(os.path.basename(path))[0]
            out = os.path.join(bill_dir, bill_name + '输出结果.xlsx')
            self.dst_var.set(out)

    def _pick_dst(self):
        path = filedialog.asksaveasfilename(
            title='选择输出文件路径',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')])
        if path:
            self.dst_var.set(path)

    def _log(self, msg):
        """线程安全写日志"""
        def _write():
            self.log_box.configure(state='normal')
            self.log_box.insert('end', msg + '\n')
            self.log_box.see('end')
            self.log_box.configure(state='disabled')
        self.after(0, _write)

    def _set_progress(self, val):
        self.after(0, lambda: self.progress.configure(value=val))

    def _run(self):
        tpl = self.tpl_var.get().strip()
        bill = self.bill_var.get().strip()
        dst = self.dst_var.get().strip()

        if not tpl:
            messagebox.showwarning('提示', '请先选择台账模板文件！')
            return
        if not os.path.exists(tpl):
            messagebox.showwarning('提示', f'模板文件不存在：\n{tpl}')
            return
        if not bill:
            messagebox.showwarning('提示', '请先选择对账单文件！')
            return
        if not dst:
            messagebox.showwarning('提示', '请先指定输出文件路径！')
            return

        self.run_btn.configure(state='disabled', text='⏳ 处理中...')
        self.progress['value'] = 0

        def task():
            try:
                fill_ledger(tpl, bill, dst, self._log, self._set_progress)
                self.after(0, lambda: messagebox.showinfo('完成', f'✅ 填充完成！\n\n输出文件：\n{dst}'))
            except Exception as e:
                self._log(f'\n❌ 出错: {e}')
                self.after(0, lambda: messagebox.showerror('错误', f'运行出错：\n{e}'))
            finally:
                self.after(0, lambda: self.run_btn.configure(state='normal', text='▶  开始填充'))

        threading.Thread(target=task, daemon=True).start()


if __name__ == '__main__':
    app = App()
    app.mainloop()
